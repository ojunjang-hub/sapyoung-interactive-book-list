"""도서목록 내보내기 — 엑셀(.xlsx) / PDF.

정본 books.db(books_effective 뷰)에서 그 자리에서 생성하므로 항상 최신 상태다.
공개 컬럼만 포함한다(내부 지표 제외). 클라이언트가 현재 화면의 ISBN 목록을
보내면 그 순서대로, 비우면 전체(최신순)를 내보낸다.
"""
import io
from datetime import datetime
from xml.sax.saxutils import escape

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel

from database import get_conn

router = APIRouter()

# (DB 컬럼, 표시 헤더) — 공개용. 내부 지표(판매지수 등)는 제외.
EXPORT_COLUMNS = [
    ("title", "제목"),
    ("author", "저자"),
    ("publisher", "출판사"),
    ("pub_date", "출간일"),
    ("department", "부서"),
    ("category", "분야"),
    ("series", "시리즈"),
    ("price_standard", "정가"),
    ("stock_status", "재고상태"),
    ("isbn13", "ISBN"),
]

_SELECT = (
    "isbn13, title, author, publisher, pub_date, "
    "department_effective AS department, "
    "category_effective   AS category, "
    "series, price_standard, stock_status"
)

_MAX_ISBNS = 5000  # 방어용 상한(전체가 882권 규모)


class ExportRequest(BaseModel):
    isbns: list[str] | None = None


def _fetch_rows(isbns: list[str] | None) -> list[dict]:
    with get_conn() as conn:
        if isbns:
            isbns = isbns[:_MAX_ISBNS]
            ph = ",".join("?" * len(isbns))
            rows = conn.execute(
                f"SELECT {_SELECT} FROM books_effective WHERE isbn13 IN ({ph})",
                isbns,
            ).fetchall()
            by = {r["isbn13"]: dict(r) for r in rows}
            return [by[i] for i in isbns if i in by]  # 클라이언트 표시 순서 보존
        rows = conn.execute(
            f"SELECT {_SELECT} FROM books_effective ORDER BY pub_date DESC"
        ).fetchall()
        return [dict(r) for r in rows]


def _build_xlsx(rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "도서목록"

    headers = [h for _, h in EXPORT_COLUMNS]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2C3E50")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")

    for r in rows:
        ws.append([r.get(k) for k, _ in EXPORT_COLUMNS])

    widths = [42, 18, 16, 12, 12, 16, 22, 10, 10, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_pdf(rows: list[dict]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (
        LongTable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        TableStyle,
    )

    # 한글: reportlab 내장 CID 폰트(별도 폰트 파일 불필요)
    pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))
    FONT = "HYSMyeongJo-Medium"

    title_st = ParagraphStyle("title", fontName=FONT, fontSize=15, leading=19)
    meta_st = ParagraphStyle("meta", fontName=FONT, fontSize=9, leading=12,
                             textColor=colors.HexColor("#666666"))
    head_st = ParagraphStyle("head", fontName=FONT, fontSize=9, leading=11,
                             textColor=colors.white)
    cell_st = ParagraphStyle("cell", fontName=FONT, fontSize=8, leading=10)

    # 가로 A4 폭에 맞춘 컬럼 구성(제목 폭 우선)
    cols = [
        ("title", "제목", 88 * mm),
        ("author", "저자", 48 * mm),
        ("publisher", "출판사", 40 * mm),
        ("pub_date", "출간일", 22 * mm),
        ("category", "분야", 40 * mm),
        ("price_standard", "정가", 22 * mm),
    ]

    def cellp(text: str, style: ParagraphStyle) -> Paragraph:
        return Paragraph(escape("" if text is None else str(text)), style)

    table_data = [[cellp(h, head_st) for _, h, _ in cols]]
    for r in rows:
        line = []
        for key, _, _ in cols:
            v = r.get(key)
            if key == "price_standard" and v is not None:
                v = f"{int(v):,}"
            line.append(cellp(v, cell_st))
        table_data.append(line)

    col_widths = [w for _, _, w in cols]
    table = LongTable(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6F8")]),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),  # 정가 우측정렬
    ]))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=12 * mm,
        title="사회평론 도서목록",
    )
    today = datetime.now().strftime("%Y-%m-%d")
    story = [
        Paragraph("사회평론 도서목록", title_st),
        Paragraph(f"총 {len(rows):,}권 · 출력일 {today}", meta_st),
        Spacer(1, 6 * mm),
        table,
    ]
    doc.build(story)
    return buf.getvalue()


@router.post("/books")
def export_books(body: ExportRequest, format: str = Query("xlsx")):
    fmt = format.lower()
    if fmt not in ("xlsx", "pdf"):
        raise HTTPException(400, "지원하지 않는 형식입니다 (xlsx 또는 pdf)")

    rows = _fetch_rows(body.isbns)
    if not rows:
        raise HTTPException(404, "내보낼 도서가 없습니다")

    if fmt == "xlsx":
        data = _build_xlsx(rows)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        data = _build_pdf(rows)
        media = "application/pdf"

    # 파일명은 ASCII(헤더 인코딩 안전). 한글 파일명은 프런트가 a.download로 지정.
    fname = f"sapyoung_booklist_{datetime.now():%Y%m%d}.{fmt}"
    headers = {"Content-Disposition": f'attachment; filename="{fname}"'}
    return Response(content=data, media_type=media, headers=headers)
